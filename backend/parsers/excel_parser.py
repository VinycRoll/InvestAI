"""CSV and XLSX parser with robust column detection and transaction extraction.

Supports Brazilian and international financial statement formats including
dual credit/debit columns, multiple encodings, and various delimiters.
"""
import logging
import re
from datetime import datetime
from io import BytesIO, StringIO

import openpyxl
import pandas as pd

from .utils import (
    CREDIT_KEYWORDS,
    DEBIT_KEYWORDS,
    NON_AMOUNT_KEYWORDS,
    is_amount_column,
    is_credit_column,
    is_date_column,
    is_desc_column,
    is_debit_column,
    normalize_text,
    parse_date,
    parse_money,
)

logger = logging.getLogger(__name__)

# Keep legacy names for backward compatibility with tests
_parse_amount = parse_money
_parse_date = parse_date


# ---------------------------------------------------------------------------
# CSV-specific helpers
# ---------------------------------------------------------------------------

def _detect_csv_delimiter(text: str) -> str:
    """Detect CSV delimiter by counting occurrences in the first 10 lines."""
    first_lines = text.split("\n")[:10]
    comma_count = sum(line.count(",") for line in first_lines)
    semicolon_count = sum(line.count(";") for line in first_lines)
    tab_count = sum(line.count("\t") for line in first_lines)

    # Tab often means tab-separated
    if tab_count > 0 and tab_count >= comma_count and tab_count >= semicolon_count:
        return "\t"
    if semicolon_count > comma_count:
        return ";"
    return ","


def _decode_with_fallback(file_content: bytes) -> tuple[str, str]:
    """Decode bytes to string trying multiple encodings.

    Returns (decoded_text, encoding_name).
    """
    if not file_content:
        return "", "utf-8"

    # Check for BOM markers first
    if file_content[:3] == b"\xef\xbb\xbf":
        return file_content[3:].decode("utf-8"), "utf-8-sig"
    if file_content[:2] in (b"\xff\xfe", b"\xfe\xff"):
        # UTF-16 BOM — not common for CSV but handle gracefully
        for enc in ("utf-16", "utf-16-le", "utf-16-be"):
            try:
                return file_content.decode(enc), enc
            except (UnicodeDecodeError, LookupError):
                continue

    for encoding in ("utf-8", "latin-1", "cp1252", "iso-8859-1"):
        try:
            return file_content.decode(encoding), encoding
        except (UnicodeDecodeError, LookupError):
            continue

    return file_content.decode("utf-8", errors="replace"), "utf-8-fallback"


# ---------------------------------------------------------------------------
# Row-level transaction extraction (shared by CSV and XLSX)
# ---------------------------------------------------------------------------

def _extract_transactions_from_rows(rows: list[dict]) -> tuple[list[dict], dict]:
    """Extract transactions from a list of row dicts.

    Returns (transactions, diagnostics) where diagnostics contains
    columns_detected, rows_processed, rejection_counts, etc.
    """
    if not rows:
        return [], {"rows_processed": 0, "transactions_found": 0, "rejection_counts": {}}

    all_cols = list(rows[0].keys())
    date_col = None
    desc_col = None
    amount_col = None
    credit_col = None
    debit_col = None

    diagnostics = {
        "rows_processed": len(rows),
        "columns_detected": {},
        "rejection_counts": {
            "missing_date": 0,
            "missing_amount": 0,
            "zero_amount": 0,
        },
    }

    # --- Detect columns ---
    for col in all_cols:
        sample = [r.get(col) for r in rows[:20] if r.get(col) is not None]
        col_lower = col.lower().strip()

        if date_col is None and is_date_column(col, sample):
            date_col = col
        elif is_credit_column(col):
            credit_col = col
        elif is_debit_column(col):
            debit_col = col
        elif amount_col is None and is_amount_column(col, sample):
            amount_col = col
        elif desc_col is None and is_desc_column(col, sample):
            desc_col = col

    # If we have credit/debit but no single amount column, use credit+debit
    use_dual = credit_col is not None and debit_col is not None and amount_col is None

    # Fallback: scan for numeric columns if amount not found
    if amount_col is None and not use_dual:
        numeric_cols = []
        for col in all_cols:
            col_lower = col.lower().strip()
            if any(kw in col_lower for kw in NON_AMOUNT_KEYWORDS):
                continue
            sample = [r.get(col) for r in rows[:20]]
            numeric_count = 0
            has_decimal = 0
            for v in sample:
                if v is None:
                    continue
                parsed = parse_money(str(v).strip())
                if parsed is not None:
                    numeric_count += 1
                    s = str(v).strip()
                    if "," in s or ("." in s and s.count(".") == 1 and len(s.split(".")[-1]) == 2):
                        has_decimal += 1
            if numeric_count >= max(1, len(sample) * 0.4):
                numeric_cols.append((col, has_decimal))
        if numeric_cols:
            numeric_cols.sort(key=lambda x: x[1], reverse=True)
            amount_col = numeric_cols[0][0]

    # Fallback: pick first non-date/non-amount column as description
    if desc_col is None:
        for col in all_cols:
            if col != date_col and col != amount_col and col != credit_col and col != debit_col:
                desc_col = col
                break

    diagnostics["columns_detected"] = {
        "date": date_col,
        "description": desc_col,
        "amount": amount_col,
        "credit": credit_col,
        "debit": debit_col,
    }

    logger.info(
        "Column detection: date=%s, desc=%s, amount=%s, credit=%s, debit=%s",
        date_col, desc_col, amount_col, credit_col, debit_col,
    )

    if not use_dual and amount_col is None:
        logger.warning("No amount column detected — returning empty transaction list")
        diagnostics["rejection_counts"]["missing_amount"] = len(rows)
        return [], diagnostics

    # --- Extract transactions ---
    transactions = []
    for row in rows:
        # Parse date
        raw_date = row.get(date_col) if date_col else None
        parsed_date = parse_date(str(raw_date)) if raw_date is not None else ""
        if not parsed_date:
            diagnostics["rejection_counts"]["missing_date"] += 1
            continue

        # Parse description
        raw_desc = row.get(desc_col) if desc_col else ""
        description = normalize_text(str(raw_desc)) if raw_desc is not None else ""

        if use_dual:
            # Dual credit/debit columns
            raw_credit = row.get(credit_col) if credit_col else None
            raw_debit = row.get(debit_col) if debit_col else None
            credit_val = parse_money(str(raw_credit)) if raw_credit is not None else None
            debit_val = parse_money(str(raw_debit)) if raw_debit is not None else None

            if credit_val is None and debit_val is None:
                diagnostics["rejection_counts"]["missing_amount"] += 1
                continue
            if credit_val is None:
                credit_val = 0.0
            if debit_val is None:
                debit_val = 0.0

            # Credit is positive, debit is negative
            amount = credit_val - debit_val
            if amount == 0:
                diagnostics["rejection_counts"]["zero_amount"] += 1
                continue
        else:
            # Single amount column
            raw_amount = row.get(amount_col) if amount_col else None
            if raw_amount is None:
                diagnostics["rejection_counts"]["missing_amount"] += 1
                continue
            amount = parse_money(str(raw_amount))
            if amount is None:
                diagnostics["rejection_counts"]["missing_amount"] += 1
                continue

        transactions.append({
            "date": parsed_date,
            "description": description,
            "amount": round(amount, 2),
        })

    diagnostics["transactions_found"] = len(transactions)

    logger.info(
        "Extracted %d transactions from %d rows (rejected: %s)",
        len(transactions), len(rows), diagnostics["rejection_counts"],
    )

    return transactions, diagnostics


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

def parse_excel(file_content: bytes, filename: str = "") -> dict:
    """Parse an Excel XLSX/XLS file and extract transactions."""
    logger.info("Starting XLSX parse for filename=%s, size=%d bytes", filename, len(file_content))
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        sheets_data = {}
        total_rows = 0
        sheet_names = list(wb.sheetnames)

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = []
            headers = None

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h else f"col_{j}" for j, h in enumerate(row)]
                else:
                    row_dict = {headers[j]: row[j] for j in range(len(row)) if j < len(headers)}
                    rows.append(row_dict)

            sheets_data[sheet_name] = rows
            total_rows += len(rows)

        wb.close()

        all_rows = []
        for sheet_name, rows in sheets_data.items():
            for row in rows:
                row["_sheet"] = sheet_name
                all_rows.append(row)

        transactions, tx_diag = _extract_transactions_from_rows(all_rows)

        total_income = sum(t["amount"] for t in transactions if t["amount"] > 0)
        total_expenses = sum(abs(t["amount"]) for t in transactions if t["amount"] < 0)

        df = pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
        summary_stats = {}
        if not df.empty:
            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            for col in numeric_cols:
                summary_stats[col] = {
                    "sum": round(float(df[col].sum()), 2),
                    "mean": round(float(df[col].mean()), 2),
                    "min": round(float(df[col].min()), 2),
                    "max": round(float(df[col].max()), 2),
                }

        diagnostics = {
            "rows_processed": tx_diag.get("rows_processed", 0),
            "transactions_found": tx_diag.get("transactions_found", 0),
            "columns_detected": tx_diag.get("columns_detected", {}),
            "rejection_counts": tx_diag.get("rejection_counts", {}),
        }

        logger.info(
            "XLSX parse complete: %d sheets, %d rows, %d transactions",
            len(sheets_data), total_rows, len(transactions),
        )

        return {
            "type": "XLSX",
            "sheets": sheet_names,
            "total_rows": total_rows,
            "sheets_data": {k: v[:100] for k, v in sheets_data.items()},
            "summary_stats": summary_stats,
            "columns": list(df.columns) if not df.empty else [],
            "transactions": transactions,
            "total_transactions": len(transactions),
            "total_income": round(total_income, 2),
            "total_expenses": round(total_expenses, 2),
            "summary": (
                f"Planilha Excel com {len(sheets_data)} aba(s) e {total_rows} linhas. "
                f"Transações: {len(transactions)}. "
                f"Receitas: R$ {total_income:,.2f}. Despesas: R$ {total_expenses:,.2f}."
            ),
            "_diagnostics": diagnostics,
        }

    except Exception as e:
        logger.exception("XLSX parse failed")
        return {
            "type": "XLSX",
            "error": str(e),
            "transactions": [],
            "sheets_data": {},
            "summary": f"Erro ao processar Excel: {e}",
        }


# ---------------------------------------------------------------------------
# CSV parser
# ---------------------------------------------------------------------------

def parse_csv(file_content: bytes, filename: str = "") -> dict:
    """Parse a CSV file and extract transactions.

    Supports multiple encodings (UTF-8, UTF-8 BOM, CP1252, Latin-1),
    delimiters (comma, semicolon, tab), and both single amount column
    and dual credit/debit column formats.
    """
    logger.info("Starting CSV parse for filename=%s, size=%d bytes", filename, len(file_content))

    try:
        if not file_content or not file_content.strip():
            return {
                "type": "CSV",
                "transactions": [],
                "data": [],
                "summary": "Arquivo CSV vazio.",
                "_diagnostics": {"rows_processed": 0, "transactions_found": 0, "rejection_counts": {}},
            }

        text, encoding = _decode_with_fallback(file_content)
        logger.info("Detected encoding: %s", encoding)

        delimiter = _detect_csv_delimiter(text)
        logger.info("Detected delimiter: %r", delimiter)

        df = pd.read_csv(StringIO(text), sep=delimiter, on_bad_lines="skip", dtype=str)

        if df.empty or len(df.columns) < 2:
            return {
                "type": "CSV",
                "columns": list(df.columns) if not df.empty else [],
                "total_rows": 0,
                "data": [],
                "transactions": [],
                "summary": "Arquivo CSV sem dados tabulares válidos.",
                "_diagnostics": {"rows_processed": 0, "transactions_found": 0, "rejection_counts": {}},
            }

        # Normalize column names: strip whitespace
        df.columns = [c.strip() for c in df.columns]

        rows = df.head(500).to_dict(orient="records")
        transactions, tx_diag = _extract_transactions_from_rows(rows)

        total_income = sum(t["amount"] for t in transactions if t["amount"] > 0)
        total_expenses = sum(abs(t["amount"]) for t in transactions if t["amount"] < 0)

        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
        summary_stats = {}
        for col in numeric_cols:
            summary_stats[col] = {
                "sum": round(float(df[col].sum()), 2),
                "mean": round(float(df[col].mean()), 2),
                "min": round(float(df[col].min()), 2),
                "max": round(float(df[col].max()), 2),
            }

        diagnostics = {
            "rows_processed": tx_diag.get("rows_processed", 0),
            "transactions_found": tx_diag.get("transactions_found", 0),
            "columns_detected": tx_diag.get("columns_detected", {}),
            "rejection_counts": tx_diag.get("rejection_counts", {}),
            "encoding": encoding,
            "delimiter": delimiter,
        }

        logger.info(
            "CSV parse complete: %d rows, %d columns, %d transactions",
            len(df), len(df.columns), len(transactions),
        )

        return {
            "type": "CSV",
            "columns": list(df.columns),
            "total_rows": len(df),
            "data": rows[:200],
            "summary_stats": summary_stats,
            "transactions": transactions,
            "total_transactions": len(transactions),
            "total_income": round(total_income, 2),
            "total_expenses": round(total_expenses, 2),
            "summary": (
                f"Planilha CSV com {len(df)} linhas e {len(df.columns)} colunas. "
                f"Transações: {len(transactions)}. "
                f"Receitas: R$ {total_income:,.2f}. Despesas: R$ {total_expenses:,.2f}."
            ),
            "_diagnostics": diagnostics,
        }

    except Exception as e:
        logger.exception("CSV parse failed")
        return {
            "type": "CSV",
            "error": str(e),
            "transactions": [],
            "data": [],
            "summary": f"Erro ao processar CSV: {e}",
            "_diagnostics": {"rows_processed": 0, "transactions_found": 0, "rejection_counts": {}},
        }
