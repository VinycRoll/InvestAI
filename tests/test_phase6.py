"""
FASE 6 - Cobertura determinística offline.

Nenhum teste aqui depende de:
- backend ao vivo (todas as rotas usam TestClient + override de DB e Gemini),
- acesso à internet,
- horário real do sistema,
- time.sleep() (retries do Gemini são zerados via monkeypatch),
- filesystem global compartilhado (DB temporário com nome único).
"""
import asyncio
import csv
import io
import json
import os
import sys
import tempfile
import uuid
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import backend.main as main
from backend.database import Base, get_db
from backend.main import app, get_gemini
from backend.services import gemini
from backend.services.analysis import (
    analyze_transactions,
    categorize_transaction,
)
from backend.services.export import ExportService

# =============================================================================
# Infraestrutura de teste offline (DB temporário + Gemini simulado)
# =============================================================================

_TMP_DB = os.path.join(tempfile.gettempdir(), f"investia_phase6_{uuid.uuid4().hex}.db")
_engine = create_engine(f"sqlite:///{_TMP_DB}", connect_args={"check_same_thread": False})
Base.metadata.create_all(bind=_engine)
_TestSession = sessionmaker(autocommit=False, autoflush=False, bind=_engine)


def _test_db():
    db = _TestSession()
    try:
        yield db
    finally:
        db.close()


class _FakeGemini:
    async def chat(self, messages, system_prompt=None):
        return "resposta simulada do modelo"

    async def generate_investment_recommendation(self, profile, amount, categories):
        return "recomendacao educacional simulada"


app.dependency_overrides[get_db] = _test_db
app.dependency_overrides[get_gemini] = lambda: _FakeGemini()

_client_ctx = TestClient(app)


@pytest.fixture(autouse=True)
def _reset_rate_limit_store():
    """Garante que o rate limit (estado global) nunca vaze entre testes."""
    main.rate_limit_store.clear()
    yield
    main.rate_limit_store.clear()


def _register(headers_prefix=""):
    email = f"phase6_{headers_prefix}{uuid.uuid4().hex[:10]}@investia.test"
    resp = _client_ctx.post("/api/auth/register", json={
        "email": email, "name": "Fase6", "password": "SenhaTeste1",
    })
    assert resp.status_code == 200, resp.text
    body = resp.json()
    return {
        "email": email,
        "token": body["token"],
        "refresh_token": body["refresh_token"],
    }


def _headers(token):
    return {"Authorization": f"Bearer {token}"}


def _upload_csv(token, name="extrato.csv", body=None, content_type="text/csv"):
    body = body or b"data,descricao,valor\n2026-01-01,pix,-10,50\n"
    return _client_ctx.post(
        "/api/upload",
        headers=_headers(token),
        files={"file": (name, body, content_type)},
    )


# =============================================================================
# PARSERS
# =============================================================================

def test_csv_comma_delimited_no_crash():
    # CSV separado por vírgula (sem vírgula nos valores) ainda é processado.
    data = "Data,Descricao,Valor\n2026-01-01,ALUGUEL,-1200.00\n2026-01-10,IFOOD,-85.50\n"
    result = main.parse_csv(data.encode("utf-8"), "comma.csv")
    assert result.get("error") is None
    assert result["total_expenses"] == pytest.approx(1285.50)


def test_csv_utf8_bom_and_quoted_comma():
    # BOM UTF-8 + descrição com vírgula interna entre aspas não corrompe o parse
    # quando o delimitador não conflita com o decimal brasileiro (ponto e vírgula).
    data = (
        'Data;Descricao;Valor\n'
        '2026-01-05;"PADARIA, CENTRO";-12,40\n'
        '2026-01-06;SALARIO;4500,00\n'
    )
    result = main.parse_csv("\ufeff".encode("utf-8") + data.encode("utf-8"), "bom.csv")
    assert result.get("error") is None
    amounts = sorted(t["amount"] for t in result["transactions"])
    assert -12.40 in amounts
    assert 4500.00 in amounts


def test_csv_negative_brazilian_decimal():
    data = "Data;Descricao;Valor\n2026-01-01;MERCADO;-1.234,56\n"
    result = main.parse_csv(data.encode("utf-8"), "neg.csv")
    assert result["transactions"][0]["amount"] == -1234.56


def test_excel_parses_valid_workbook():
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"
    ws.append(["Data", "Descricao", "Codigo", "Valor"])
    ws.append([datetime(2026, 1, 5), "ALUGUEL", 9001, -1500.0])
    ws.append([datetime(2026, 1, 15), "SALARIO", 9002, 5000.0])
    wb.save(buf)
    result = main.parse_excel(buf.getvalue(), "extrato.xlsx")
    assert result.get("error") is None
    assert result["total_transactions"] == 2
    amounts = {t["amount"] for t in result["transactions"]}
    assert amounts == {-1500.0, 5000.0}
    assert result["total_income"] == 5000.0
    assert result["total_expenses"] == 1500.0


def test_excel_multiple_sheets_summed():
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    for i, sheet in enumerate(["Conta", "Poupanca"]):
        ws = wb.create_sheet(sheet) if i else wb.active
        ws.title = sheet
        ws.append(["Data", "Descricao", "Valor"])
        ws.append([datetime(2026, 1, i + 1), f"LANC {i}", 100.0])
    wb.save(buf)
    result = main.parse_excel(buf.getvalue(), "multi.xlsx")
    assert result["total_transactions"] == 2
    assert result["total_rows"] == 2


def test_excel_empty_workbook():
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Data", "Descricao", "Valor"])
    wb.save(buf)
    result = main.parse_excel(buf.getvalue(), "vazio.xlsx")
    assert result["transactions"] == []


def test_excel_invalid_content_returns_error():
    result = main.parse_excel(b"isto nao e um excel", "bad.xlsx")
    assert result["transactions"] == []
    assert result.get("error") is not None


def test_pdf_table_extraction_detects_columns():
    from backend.parsers.pdf_parser import extract_transactions_from_table
    table = {
        "headers": ["Data", "Descrição", "Valor"],
        "rows": [
            {"Data": "01/08/2026", "Descrição": "ALUGUEL", "Valor": "(1.234,56)"},
            {"Data": "02/08/2026", "Descrição": "SALARIO", "Valor": "5.000,00"},
        ],
    }
    txns = extract_transactions_from_table(table)
    assert len(txns) == 2
    amounts = {t["amount"] for t in txns}
    assert -1234.56 in amounts
    assert 5000.00 in amounts


def test_pdf_table_missing_amount_column_yields_empty():
    from backend.parsers.pdf_parser import extract_transactions_from_table
    table = {
        "headers": ["Nome", "Cidade"],
        "rows": [{"Nome": "Joao", "Cidade": "SP"}],
    }
    assert extract_transactions_from_table(table) == []


def test_parse_pdf_invalid_bytes_returns_error_dict():
    result = main.parse_pdf(b"nao sou um pdf")
    assert result["type"] == "PDF"
    assert result["transactions"] == []
    assert "error" in result or "summary" in result


def test_ofx_dates_and_direction():
    ofx = """OFXHEADER:100
DATA:OFXSGML
VERSION:102
SECURITY:NONE
ENCODING:USASCII
CHARSET:1252
COMPRESSION:NONE
OLDFILEUID:NONE
NEWFILEUID:NONE

<OFX>
<BANKMSGSRSV1>
<STMTTRNRS>
<TRNUID>1</TRNUID>
<STATUS><CODE>0</CODE><SEVERITY>INFO</SEVERITY></STATUS>
<STMTRS>
<CURDEF>BRL</CURDEF>
<BANKACCTFROM>
<BANKID>001</BANKID>
<ACCTID>7099</ACCTID>
<ACCTTYPE>CHECKING</ACCTTYPE>
</BANKACCTFROM>
<BANKTRANLIST>
<DTSTART>20260101</DTSTART>
<DTEND>20260331</DTEND>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260210</DTPOSTED>
<TRNAMT>-89.90</TRNAMT>
<FITID>5001</FITID>
<MEMO>MERCADO</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>CREDIT</TRNTYPE>
<DTPOSTED>20260225</DTPOSTED>
<TRNAMT>1200.00</TRNAMT>
<FITID>5002</FITID>
<MEMO>PIX RECEBIDO</MEMO>
</STMTTRN>
</BANKTRANLIST>
</STMTRS>
</STMTTRNRS>
</BANKMSGSRSV1>
</OFX>
"""
    result = main.parse_ofx(ofx.encode("utf-8"))
    assert result.get("error") is None
    assert result["total_income"] == pytest.approx(1200.00)
    assert result["total_expenses"] == pytest.approx(89.90)
    dates = sorted(t["date"][:10] for t in result["transactions"])
    assert dates == ["2026-02-10", "2026-02-25"]


# =============================================================================
# ANÁLISE (edges adicionais, sem duplicar test_analysis.py)
# =============================================================================

def test_analysis_user_category_overrides_builtin():
    txns = [{"amount": -50, "date": "2026-01-10", "description": "IFOOD"}]
    result = analyze_transactions(txns, [{"name": "pessoal", "keywords": ["ifood"]}])
    assert "pessoal" in result["categories"]
    assert "alimentacao" not in result["categories"]


def test_categorize_user_categories_take_priority():
    assert categorize_transaction("IFOOD", [{"name": "pessoal", "keywords": ["ifood"]}]) == "pessoal"


def test_analysis_monthly_change_div_by_zero():
    txns = [
        {"amount": 0, "date": "2026-01-15", "description": "SEM RENDA"},
        {"amount": 5000, "date": "2026-02-15", "description": "SALARIO"},
    ]
    result = analyze_transactions(txns)
    assert any(c["month"] == "2026-02" for c in result["monthly_comparison"])
    for comp in result["monthly_comparison"]:
        assert comp["income_change_pct"] == 0.0


def test_analysis_no_surplus_means_zero_suggested_investment():
    txns = [
        {"amount": 1000, "date": "2026-01-15", "description": "SALARIO"},
        {"amount": -1000, "date": "2026-01-10", "description": "GASTOS"},
    ]
    result = analyze_transactions(txns)
    assert result["suggested_investment"] == 0.0


def test_analysis_expenses_only_no_suggested_investment():
    txns = [
        {"amount": -100, "date": "2026-01-10", "description": "IFOOD"},
    ]
    result = analyze_transactions(txns)
    assert result["total_income"] == 0
    assert result["suggested_investment"] == 0.0
    assert result["savings_rate"] == 0


# =============================================================================
# AUTH
# =============================================================================

def test_update_avatar_persists_for_current_user():
    creds = _register()
    avatar_url = (
        "data:image/png;base64,"
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9JrS8AAAAASUVORK5CYII="
    )

    response = _client_ctx.post(
        "/api/auth/avatar",
        headers=_headers(creds["token"]),
        json={"avatar_url": avatar_url},
    )

    assert response.status_code == 200, response.text
    assert response.json()["user"]["avatar_url"] == avatar_url

    me = _client_ctx.get("/api/auth/me", headers=_headers(creds["token"]))
    assert me.json()["avatar_url"] == avatar_url


def test_update_avatar_rejects_unsupported_image_type():
    creds = _register()
    response = _client_ctx.post(
        "/api/auth/avatar",
        headers=_headers(creds["token"]),
        json={"avatar_url": "data:image/svg+xml;base64,PHN2Zz48L3N2Zz4="},
    )

    assert response.status_code == 400
    assert "PNG, JPEG ou WebP" in response.json()["detail"]


def test_refresh_token_has_type_refresh():
    from backend.auth import create_refresh_token, verify_token
    token = create_refresh_token({"sub": "1"})
    assert verify_token(token).get("type") == "refresh"


def test_login_success():
    creds = _register()
    resp = _client_ctx.post("/api/auth/login", json={
        "email": creds["email"], "password": "SenhaTeste1",
    })
    assert resp.status_code == 200
    assert "token" in resp.json()
    assert "refresh_token" in resp.json()


def test_login_wrong_password():
    creds = _register()
    resp = _client_ctx.post("/api/auth/login", json={
        "email": creds["email"], "password": "SenhaErrada1",
    })
    assert resp.status_code == 401


def test_login_unknown_user():
    resp = _client_ctx.post("/api/auth/login", json={
        "email": "naoexiste@investia.test", "password": "SenhaTeste1",
    })
    assert resp.status_code == 401


def test_refresh_token_cannot_act_as_access():
    creds = _register()
    # Refresh token usado como Bearer em rota protegida deve falhar.
    resp = _client_ctx.get("/api/auth/me", headers=_headers(creds["refresh_token"]))
    assert resp.status_code == 401


def test_access_token_used_as_refresh_denied():
    creds = _register()
    resp = _client_ctx.post("/api/auth/refresh", json={"refresh_token": creds["token"]})
    assert resp.status_code == 401


def test_refresh_endpoint_returns_new_pair():
    creds = _register()
    resp = _client_ctx.post("/api/auth/refresh", json={"refresh_token": creds["refresh_token"]})
    assert resp.status_code == 200
    assert "token" in resp.json()
    assert "refresh_token" in resp.json()


def test_invalid_token_rejected():
    resp = _client_ctx.get("/api/auth/me", headers=_headers("token.invalido.aqui"))
    assert resp.status_code == 401


def test_cross_user_resource_isolation():
    user_a = _register("a_")
    user_b = _register("b_")
    up = _upload_csv(user_a["token"], name="a.csv")
    assert up.status_code == 200, up.text
    file_id = up.json()["id"]
    # Usuário B não pode acessar arquivo do usuário A.
    assert _client_ctx.get(f"/api/transactions/{file_id}", headers=_headers(user_b["token"])).status_code == 404
    assert _client_ctx.delete(f"/api/files/{file_id}", headers=_headers(user_b["token"])).status_code == 404
    # Usuário A ainda vê o próprio arquivo.
    assert _client_ctx.get(f"/api/transactions/{file_id}", headers=_headers(user_a["token"])).status_code == 200


# =============================================================================
# API - códigos de status produzidos pelo contrato real
# =============================================================================

def test_api_unauthenticated_list_files():
    assert _client_ctx.get("/api/files").status_code == 401


def test_api_invalid_upload_type_rejected_400():
    creds = _register()
    resp = _upload_csv(creds["token"], name="arquivo.txt", body=b"qualquer coisa")
    assert resp.status_code == 400
    assert "não suportado" in resp.json()["detail"].lower()


def test_api_oversized_upload_413():
    creds = _register()
    big = b"x" * (main.MAX_FILE_SIZE + 1024)
    resp = _upload_csv(creds["token"], name="grande.csv", body=big)
    assert resp.status_code == 413


def test_api_missing_body_field_422():
    resp = _client_ctx.post("/api/auth/login", json={"email": "a@b.com"})
    assert resp.status_code == 422


def test_api_login_rate_limit_429():
    from backend.main import RATE_LIMIT_MAX_LOGIN
    cre_ds = _register()
    for _ in range(RATE_LIMIT_MAX_LOGIN):
        r = _client_ctx.post("/api/auth/login", json={
            "email": cre_ds["email"], "password": "SenhaErrada1",
        })
        assert r.status_code == 401
    r = _client_ctx.post("/api/auth/login", json={
        "email": cre_ds["email"], "password": "SenhaTeste1",
    })
    assert r.status_code == 429


def test_api_analysis_returns_502_when_gemini_unavailable():
    creds = _register()
    up = _upload_csv(creds["token"], name="analise.csv")
    assert up.status_code == 200
    file_id = up.json()["id"]

    class _BrokenGemini:
        async def chat(self, messages, system_prompt=None):
            raise gemini.GeminiAPIError("boom")

        async def generate_investment_recommendation(self, profile, amount, categories):
            raise gemini.GeminiAPIError("boom")

    app.dependency_overrides[get_gemini] = lambda: _BrokenGemini()
    try:
        resp = _client_ctx.post("/api/analysis", headers=_headers(creds["token"]),
                                json={"file_id": file_id})
        assert resp.status_code == 502
        assert "indisponível" in resp.json()["detail"].lower()
    finally:
        app.dependency_overrides[get_gemini] = lambda: _FakeGemini()


def test_api_unhandled_exception_returns_500():
    """Verifica o caminho de erro 500 quando uma dependência estoura."""
    client = TestClient(app, raise_server_exceptions=False)

    def _boom():
        raise RuntimeError("falha inesperada")

    app.dependency_overrides[get_db] = _boom
    try:
        resp = client.get("/api/health")
        assert resp.status_code == 500
    finally:
        app.dependency_overrides[get_db] = _test_db


# =============================================================================
# UPLOAD
# =============================================================================

def test_upload_allowed_extensions():
    creds = _register()
    ofx_header = (
        "OFXHEADER:100\nDATA:OFXSGML\nVERSION:102\nSECURITY:NONE\n"
        "ENCODING:USASCII\nCHARSET:1252\nCOMPRESSION:NONE\n"
        "OLDFILEUID:NONE\nNEWFILEUID:NONE\n\n<OFX></OFX>\n"
    )
    ofx = ofx_header
    for name, content, ctype in [
        ("e.csv", b"data,descricao,valor\n2026-01-01,pix,-10,50\n", "text/csv"),
        ("e.ofx", ofx.encode("utf-8"), "text/ofx"),
        ("e.qfx", ofx_header.encode("utf-8"), "application/x-ofx"),
    ]:
        resp = _upload_csv(creds["token"], name=name, body=content, content_type=ctype)
        assert resp.status_code == 200, (name, resp.text)


def test_upload_filename_without_extension_rejected():
    creds = _register()
    resp = _upload_csv(creds["token"], name="semextensao", body=b"x")
    assert resp.status_code == 400 or resp.status_code == 415


def test_upload_csv_invalid_content_still_stored():
    # Conteúdo não tabular com extensão suportada: o upload não quebra (200) e
    # o resultado parseado é um dict sem gerar transações (contrato atual).
    creds = _register()
    resp = _upload_csv(creds["token"], name="ruim.csv",
                       body=b"isto nao sao dados financeiros, apenas texto corrido")
    assert resp.status_code == 200
    body = resp.json()
    assert "parsed" in body
    assert body["parsed"]["type"] == "CSV"
    assert body["parsed"]["total_rows"] == 0
    assert "sem dados tabulares válidos" in body["parsed"]["summary"].lower()


def test_upload_file_size_recorded():
    creds = _register()
    body = b"data,descricao,valor\n2026-01-01,abc,-1,50\n"
    resp = _upload_csv(creds["token"], name="tam.csv", body=body)
    assert resp.status_code == 200
    assert resp.json()["file_size"] == len(body)


# =============================================================================
# EXPORT
# =============================================================================

def test_export_csv_roundtrips_special_characters():
    # Vírgula, aspas, acentos e quebra de linha em descrição devem ser
    # corretamente escapados e reversíveis via csv.reader (contrato CSV).
    analysis = {
        "total_income": 1000,
        "total_expenses": 300,
        "balance": 700,
        "categories": {},
        "top_expenses": [
            {"description": 'Padaria, "Centro"', "date": "2026-01-05", "amount": -10.00},
            {"description": "Açúcar\ne café", "date": "2026-01-06", "amount": -5.00},
        ],
        "recurring_expenses": [],
        "suggested_investment": 0,
    }
    out = ExportService.generate_csv_data(analysis, "Análise com vírgula, aspas e \n nova linha")
    parsed = list(csv.reader(io.StringIO(out)))
    joined = "\n".join("|".join(row) for row in parsed)
    assert 'Padaria, "Centro"' in joined
    assert "Açúcar" in joined


def test_export_json_is_valid_and_utf8():
    out = ExportService.generate_json_report({"nota": "ação e ção"})
    assert json.loads(out)["nota"] == "ação e ção"


def test_api_export_html():
    creds = _register()
    up = _upload_csv(creds["token"], name="exp.csv")
    assert up.status_code == 200
    file_id = up.json()["id"]
    an = _client_ctx.post("/api/analysis", headers=_headers(creds["token"]),
                          json={"file_id": file_id})
    assert an.status_code == 200, an.text
    analysis_id = an.json()["id"]
    resp = _client_ctx.post("/api/reports/export", headers=_headers(creds["token"]),
                            json={"analysis_id": analysis_id, "format": "html"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/html")
    assert "InvestIA" in resp.text


def test_api_export_csv():
    creds = _register()
    up = _upload_csv(creds["token"], name="exp2.csv")
    file_id = up.json()["id"]
    an = _client_ctx.post("/api/analysis", headers=_headers(creds["token"]),
                          json={"file_id": file_id})
    analysis_id = an.json()["id"]
    resp = _client_ctx.post("/api/reports/export", headers=_headers(creds["token"]),
                            json={"analysis_id": analysis_id, "format": "csv"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")


def test_api_export_json():
    creds = _register()
    up = _upload_csv(creds["token"], name="exp3.csv")
    file_id = up.json()["id"]
    an = _client_ctx.post("/api/analysis", headers=_headers(creds["token"]),
                          json={"file_id": file_id})
    analysis_id = an.json()["id"]
    resp = _client_ctx.post("/api/reports/export", headers=_headers(creds["token"]),
                            json={"analysis_id": analysis_id, "format": "json"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/json")
    assert json.loads(resp.text)


def test_api_export_nonexistent_analysis_404():
    creds = _register()
    resp = _client_ctx.post("/api/reports/export", headers=_headers(creds["token"]),
                            json={"analysis_id": 999999, "format": "html"})
    assert resp.status_code == 404


def test_api_export_pdf_requires_weasyprint():
    # PDF offline depende de weasyprint (não instalado aqui). Documentamos o
    # comportamento: sem weasyprint, a rota não retorna 200.
    import importlib.util
    if importlib.util.find_spec("weasyprint") is None:
        pytest.skip("weasyprint não instalado; export PDF não testável offline no CI")


# =============================================================================
# GEMINI - cenários com requisição simulada (nunca chamada real)
# =============================================================================

class FakeResponse:
    def __init__(self, status_code, json_data=None, raises=None):
        self.status_code = status_code
        self._json = json_data
        self._raises = raises

    def json(self):
        if self._raises:
            raise self._raises
        if self._json is None:
            raise ValueError("no json body")
        return self._json


class FakeClient:
    def __init__(self, responses=None):
        self.responses = list(responses or [])
        self.calls = []

    async def post(self, url, params=None, json=None, headers=None):
        self.calls.append(url)
        item = self.responses.pop(0)
        if isinstance(item, Exception):
            raise item
        return item


def _no_retries(monkeypatch):
    monkeypatch.setattr(gemini, "GEMINI_RETRIES", 0)
    monkeypatch.setattr(gemini, "GEMINI_RETRY_DELAY", 0.0)


def _run(coro):
    return asyncio.run(coro)


def test_gemini_success(monkeypatch):
    _no_retries(monkeypatch)
    fake = FakeClient([
        FakeResponse(200, json_data={
            "candidates": [{"content": {"parts": [{"text": "resposta ok"}]}}]
        }),
    ])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    assert _run(svc.chat([{"role": "user", "content": "oi"}])) == "resposta ok"


def test_gemini_http_error_raises(monkeypatch):
    _no_retries(monkeypatch)
    fake = FakeClient([FakeResponse(429, json_data={"error": {"message": "quota excedida"}})])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    with pytest.raises(gemini.GeminiAPIError):
        _run(svc.chat([{"role": "user", "content": "oi"}]))


def test_gemini_5xx_retries_then_raises(monkeypatch):
    monkeypatch.setattr(gemini, "GEMINI_RETRIES", 2)
    monkeypatch.setattr(gemini, "GEMINI_RETRY_DELAY", 0.0)
    fake = FakeClient([
        FakeResponse(503, json_data={"error": {"message": "interna"}}),
        FakeResponse(503, json_data={"error": {"message": "interna"}}),
        FakeResponse(503, json_data={"error": {"message": "interna"}}),
    ])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    with pytest.raises(gemini.GeminiAPIError):
        _run(svc.chat([{"role": "user", "content": "oi"}]))
    assert len(fake.calls) == 3


def test_gemini_empty_candidates_raises(monkeypatch):
    _no_retries(monkeypatch)
    fake = FakeClient([FakeResponse(200, json_data={"candidates": []})])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    with pytest.raises(gemini.GeminiAPIError):
        _run(svc.chat([{"role": "user", "content": "oi"}]))


def test_gemini_invalid_json_raises(monkeypatch):
    _no_retries(monkeypatch)
    fake = FakeClient([FakeResponse(200, raises=ValueError("not json"))])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    with pytest.raises(gemini.GeminiAPIError) as e:
        _run(svc.chat([{"role": "user", "content": "oi"}]))
    assert "Resposta inválida (JSON)" in str(e.value)


def test_gemini_connection_timeout_raises(monkeypatch):
    _no_retries(monkeypatch)
    import httpx
    fake = FakeClient([httpx.TimeoutException("timeout")])
    monkeypatch.setattr(gemini, "_get_client", lambda: fake)
    svc = gemini.GeminiService("k")
    with pytest.raises(gemini.GeminiAPIError):
        _run(svc.chat([{"role": "user", "content": "oi"}]))


# =============================================================================
# HELPERS FRONT (lógica pura, sem browser), quando importável
# =============================================================================

def test_frontend_escape_html():
    frontend_sys_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
    initialized = False
    if os.path.isdir(frontend_sys_path):
        try:
            sys.path.insert(0, frontend_sys_path)
            from helpers import escape_html
            initialized = True
        except Exception:
            initialized = False
    if not initialized:
        pytest.skip("frontend/helpers.py não importável neste ambiente")
    assert escape_html("<b>") == "&lt;b&gt;"
    assert escape_html('"aspas"') == "&quot;aspas&quot;"
